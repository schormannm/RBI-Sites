from openpyxl import Workbook
from openpyxl import load_workbook
import dict_digger

output = {"instanceID": {"column": 1, "heading": "Instance ID"},
          "site_name": {"column": 2, "heading": "Site Name"},
          "site_number": {"column": 3, "heading": "Site Number"},
          "inspector_name": {"column": 4, "heading": "Inspector name"},
          "region": {"column": 5, "heading": "Region"},
          "gps_location": {"column": 6, "heading": "GPS Location"},
          "date_of_inspection": {"column": 7, "heading": "Date of Inspection"},
          "date_of_next_inspection_recommended": {"column": 8, "heading": "Next inspection date"},
          "tower_owner": {"column": 9, "heading": "Tower Owner"},
          "tower_type": {"column": 10, "heading": "Tower Type"},
          "manufacturer": {"column": 11, "heading": "Manufacturer"},
          "job_number": {"column": 12, "heading": "Job Number"},
          "tower_total_height": {"column": 13, "heading": "Total height"},
          "infrastructure_contractor": {"column": 14, "heading": "Contractor"},
          "tower_design": {"column": 15, "heading": "Tower Design"},
          "fall_arrest_present": {"column": 16, "heading": "Fall Arrest?"},
          "nr_of_fall_arrest_systems": {"column": 17, "heading": "# of Fall Arrests"},
          "spine_height": {"column": 18, "heading": "Spine Height"},
          "colour": {"column": 19, "heading": "Colour"},
          "latitude": {"column": 20, "heading": "Latitude"},
          "longitude": {"column": 21, "heading": "Longitude"},
          "mast_category": {"column": 22, "heading": "Category"}
          }

output_columns_main = {"instanceID": 1,
                       "site_name": 2,
                       "site_number": 3,
                       "inspector_name": 4,
                       "region": 5,
                       "gps_location": 6,
                       "date_of_inspection": 7,
                       "date_of_next_inspection_recommended": 8,
                       "tower_owner": 9,
                       "tower_type": 10,
                       "manufacturer": 11,
                       "job_number": 12,
                       "tower_total_height": 13,
                       "infrastructure_contractor": 14,
                       "tower_design": 15,
                       "fall_arrest_present": 16,
                       "nr_of_fall_arrest_systems": 17,
                       "spine_height": 18,
                       "colour": 19,
                       "latitude": 20,
                       "longitude": 21,
                       "mast_category": 22,
                       "site_classification": 23,
                       "site_release_date": 24,
                       "as_built_available": 25,
                       "fault_description": 26,
                       "mast_engineer": 27,
                       "mast_upgraded": 28,
                       "mast_upgrade_date": 29,
                       "capacity_top": 30,
                       "capacity_10_from_top": 31
                       }

output_columns_loadings = {"instanceID": 1,
                           "site_name": 2,
                           "site_number": 3,
                           "inspector_name": 4,
                           "region": 5,
                           "loading_operator": 6,
                           "nr_of_GSM_antenna": 7,
                           "size_of_GSM_antenna": 8,
                           "height_of_GSM_antenna": 9,
                           "number_of_MW_dishes": 10,
                           "date_of_structural_approval": 11,
                           "structural_approval_for": 12

                           }

output_lookup = {"instanceID": "site.@instanceID",
                 "site_name": "site.site_group.site_name",
                 "site_number": "site.site_group.site_number",
                 "inspector_name": "site.inspector_name",
                 "region": "site.site_group.region",
                 "gps_location": "site.site_group.gps_location.#text",
                 "date_of_inspection": "site.date_of_inspection",
                 "date_of_next_inspection_recommended": "site.signature_group.date_of_next_inspection_recommended",
                 "tower_owner": "site.site_group.tower_owner",
                 "tower_type": "site.tower_type",
                 "manufacturer": "site.site_group.manufacturer",
                 "job_number": "site.site_group.job_number",
                 "tower_total_height": "site.tower_top.tower_total_height",
                 "tower_design": "site.tower_group.tower_design",
                 "fall_arrest_present": "site.fall_arrest_section_exists",
                 "nr_of_fall_arrest_systems": "site.fall_arrest_section.fall_arrest_num",
                 "spine_height": "site.site_group.spine_height",
                 "colour": "site.tower_group.external_coating_colour",
                 "infrastructure_contractor": "site.site_group.infrastructure_contractor",
                 "latitude": "site.site_group.gps_location.gps_latitude",
                 "longitude": "site.site_group.gps_location.gps_longitude",
                 "mast_category": "site.grading",
                 "site_classification": "site.manual.site_classification",
                 "site_release_date": "site.manual.site_release_date",
                 "as_built_available": "site.manual.as_built_available",
                 "fault_description": "site.manual.fault_description",
                 "mast_engineer": "site.manual.mast_engineer",
                 "mast_upgraded": "site.manual.mast_upgraded",
                 "mast_upgrade_date": "site.manual.mast_upgrade_date",
                 "capacity_top": "site.manual.capacity_top",
                 "capacity_10_from_top": "site.manual.capacity_10_from_top",
                 "loading_operator": "site.loadings_group.sorted_loading_table.sorted_loading_table-repeat.@.owner",
                 "nr_of_GSM_antenna": "site.loadings_group.summary_loading_table.&.panel",
                 "size_of_GSM_antenna": "site.loadings_group.sorted_loading_table.sorted_loading_table-repeat.@.size",
                 "height_of_GSM_antenna": "site.loadings_group.sorted_loading_table.sorted_loading_table-repeat.@.height",
                 "number_of_MW_dishes": "site.loadings_group.summary_loading_table.&.panel",
                 "date_of_structural_approval": "site.loading_group.summary_loading_table.&.date_of_structural_approval",
                 "structural_approval_for": "site.loading_group.summary_loading_table.&.structural_approval_for"
                 }


# Create a workbook and add worksheets
def save_to_workbook(sites, output_file):
    workbook = Workbook()

    create_sites_worksheet(sites, workbook, "Sites", 0, output_columns_main)
    create_loadings_worksheet(sites, workbook, "Loadings", 1, output_columns_loadings)

    # Save the file
    output_filename = output_file
    workbook.save(output_filename)
    print "Output file : " + output_filename + " saved"
    return output_filename


def create_heading_row(worksheet, output_columns):
    # Start from the first cell. Rows and columns are zero indexed.
    row = 1
    for tag, value in output_columns.iteritems():
        _ = worksheet.cell(row=row, column=value, value=tag)


def create_data_rows(sites, worksheet, output_columns):
    row = 2
    for site in sites:
        # Iterate over the data and write it out row by row.
        for tag, value in output_columns.iteritems():
            col = output_columns[tag]  # get the column from the dictionary
            tag = tag_fixer(tag, row)
            value = get_site_value(site, tag)
            # print "Tag : " + tag + " Value: " + str(value)
            _ = worksheet.cell(row=row, column=col, value=value)
            col += 1

        row += 1
        col = 1


def create_data_rows_loadings(sites, worksheet, output_columns):
    row = 2
    for site in sites:
        # Iterate over the data and write it out row by row.
        mongo_str = "site.loadings_group.sorted_loading_table.sorted_loading_table-repeat"
        args = mongo_str.split(".")
        loading_table = dict_digger.dig(site, *args)
        print "Loading table"
        print loading_table
        return
        for loading in loading_table:

            for tag, value in output_columns.iteritems():
                col = output_columns[tag]  # get the column from the dictionary
                tag = tag_fixer(tag, row)
                value = get_site_value(site, tag)
                # print "Tag : " + tag + " Value: " + str(value)
                _ = worksheet.cell(row=row, column=col, value=value)
                col += 1

            row += 1
            col = 1


def tag_fixer(tag, row):
    index = row
    new_tag = tag
    return new_tag

def create_sites_worksheet(sites, workbook, sheet_title, sheet_nr, column_def):
    worksheet = workbook.create_sheet(title=sheet_title, index=sheet_nr)

    create_heading_row(worksheet, column_def)
    create_data_rows(sites, worksheet, column_def)


def create_loadings_worksheet(sites, workbook, sheet_title, sheet_nr, column_def):
    worksheet = workbook.create_sheet(title=sheet_title, index=sheet_nr)

    create_heading_row(worksheet, column_def)
    create_data_rows(sites, worksheet, column_def)


def get_site_value(site, tag):
    value = None
    mongo_str = output_lookup[tag]

    if mongo_str:
        args = mongo_str.split(".")
        value = dict_digger.dig(site, *args)

    return value
