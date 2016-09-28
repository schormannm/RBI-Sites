import datetime
import os
import string
import json
import time
import re
import urllib2
import dateparser
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl import load_workbook
import dict_digger

from flask import Flask
from flask import render_template
from flask import make_response
from flask import request
from flask.ext.login import LoginManager
from flask.ext.login import login_required
from flask.ext.login import login_user
from flask.ext.login import logout_user
from flask.ext.login import current_user
from flask import send_file

from flask import redirect
from flask import url_for
from flask import send_from_directory

from passwordhelper import PasswordHelper
from user import User
from bitlyhelper import BitlyHelper

from forms import RegistrationForm
from forms import LoginForm
from forms import SearchForm
from forms import UpdateForm

from myutils import sanitize_string
from myutils import format_date

import config
if config.test:
    from mockdbhelper import MockDBHelper as DBHelper
else:
    from dbhelper import DBHelper


DB = DBHelper()
PH = PasswordHelper()
BH = BitlyHelper()

app = Flask(__name__)
login_manager = LoginManager(app)
app.secret_key = 'dbQsAFFNGnugJ4Zmv1YJeCj4btAofOn3/Y2/iRsgA7/UvKvEJqb29NWFBvzMzRNR1xGTb0xbJZBrW+xrcCOTn7xIJBbNoIbY88Y'

output = {"instanceID": {"column": 1, "heading": "Instance ID"},
         "site_name": {"column": 2, "heading": "Site Name"},
         "site_number": {"column": 3, "heading": "Site Number"},
         "inspector_name": {"column": 4, "heading": "Inspector name"},
         "region": {"column": 5, "heading": "Region"},
         "gps_location": {"column": 6, "heading": "GPS Location"},
         "date_of_inspection": {"column": 7, "heading": "Date of Inspection"},
         "date_of_next_inspection_recommended": {"column": 8, "heading": "Next inspection date"},
         "tower_owner": {"column": 9, "heading": "Tower Owner"},
         "tower_type": {"column": 10, "heading": "Tower Type"},
         "manufacturer": {"column": 11, "heading": "Manufacturer"},
         "job_number": {"column": 12, "heading": "Job Number"},
         "tower_total_height": {"column": 13, "heading": "Total height"},
          "infrastructure_contractor": {"column": 14, "heading": "Contractor"},
          "tower_design": {"column": 15, "heading": "Tower Design"},
          "fall_arrest_present": {"column": 16, "heading": "Fall Arrest?"},
          "nr_of_fall_arrest_systems": {"column": 17, "heading": "# of Fall Arrests"},
          "spine_height": {"column": 18, "heading": "Spine Height"},
          "colour": {"column": 19, "heading": "Colour"},
          "latitude": {"column": 20, "heading": "Latitude"},
          "longitude": {"column": 21, "heading": "Longitude"},
          "mast_category": {"column": 22, "heading": "Category"}
         }

output_columns_main = {"instanceID": 1,
                 "site_name": 2,
                 "site_number": 3,
                 "inspector_name": 4,
                 "region": 5,
                 "gps_location": 6,
                 "date_of_inspection": 7,
                 "date_of_next_inspection_recommended": 8,
                 "tower_owner": 9,
                 "tower_type": 10,
                 "manufacturer": 11,
                 "job_number": 12,
                 "tower_total_height": 13,
                  "infrastructure_contractor": 14,
                  "tower_design": 15,
                  "fall_arrest_present": 16,
                  "nr_of_fall_arrest_systems": 17,
                  "spine_height": 18,
                  "colour": 19,
                  "latitude": 20,
                  "longitude": 21,
                  "mast_category": 22,
                  "site_classification": 23,
                  "site_release_date": 24,
                  "as_built_available": 25,
                  "fault_description": 26,
                  "mast_engineer": 27,
                  "mast_upgraded": 28,
                  "mast_upgrade_date": 29,
                  "capacity_top": 30,
                  "capacity_10_from_top": 31
                 }

output_columns_loadings = {"instanceID": 1,
                 "site_name": 2,
                 "site_number": 3,
                 "inspector_name": 4,
                 "region": 5,
                  "loading_operator": 6,
                  "nr_of_GSM_antenna": 7,
                  "size_of_GSM_antenna": 8,
                  "height_of_GSM_antenna": 9,
                  "number_of_MW_dishes": 10,
                  "size_of_MW_dish": 11,
                  "height_of_MW_dish": 12,
                  "date_of_structural_approval": 13,
                  "structural_approval_for": 14

                  }

output_lookup = {"instanceID": "site.@instanceID",
                 "site_name": "site.site_group.site_name",
                 "site_number": "site.site_group.site_number",
                 "inspector_name": "site.inspector_name",
                 "region": "site.site_group.region",
                 "gps_location": "site.site_group.gps_location.#text",
                 "date_of_inspection": "site.date_of_inspection",
                 "date_of_next_inspection_recommended": "site.signature_group.date_of_next_inspection_recommended",
                 "tower_owner": "site.site_group.tower_owner",
                 "tower_type": "site.tower_type",
                 "manufacturer": "site.site_group.manufacturer",
                 "job_number": "site.site_group.job_number",
                 "tower_total_height": "site.tower_top.tower_total_height",
                 "tower_design": "site.tower_group.tower_design",
                 "fall_arrest_present": "site.fall_arrest_section_exists",
                 "nr_of_fall_arrest_systems": "site.fall_arrest_section.fall_arrest_num",
                 "spine_height": "site.site_group.spine_height",
                 "colour": "site.tower_group.external_coating_colour",
                 "infrastructure_contractor": "site.site_group.infrastructure_contractor",
                 "latitude": "site.site_group.gps_location.gps_latitude",
                 "longitude": "site.site_group.gps_location.gps_longitude",
                 "mast_category": "site.grading",
                 "site_classification": "",
                 "site_release_date": "",
                 "as_built_available": "",
                 "fault_description": "",
                 "mast_engineer": "",
                 "mast_upgraded": "",
                 "mast_upgrade_date": "",
                 "capacity_top": "",
                 "capacity_10_from_top": "",
                 "loading_operator": "site.@instanceID",
                 "nr_of_GSM_antenna": "site.@instanceID",
                 "size_of_GSM_antenna": "site.@instanceID",
                 "height_of_GSM_antenna": "site.@instanceID",
                 "number_of_MW_dishes": "site.@instanceID",
                 "size_of_MW_dish": "site.@instanceID",
                 "height_of_MW_dish": "site.@instanceID",
                 "date_of_structural_approval": "site.@instanceID",
                 "structural_approval_for": "site.@instanceID"

                 }


@app.route("/")
def home():
    return render_template("home.html", loginform=LoginForm(), registrationform=RegistrationForm())


@app.route("/new-user")
def register_new_user():
    now = datetime.datetime.now()
    return render_template("register_new_user.html",  registrationform=RegistrationForm())


@login_manager.user_loader
def load_user(user_id):
    user_password = DB.get_user(user_id)
    if user_password:
        return User(user_id)


@app.route("/login", methods=["POST"])
def login():
    form = LoginForm(request.form)
    if form.validate():
        stored_user = DB.get_user(form.loginemail.data)
        if stored_user and PH.validate_password(form.loginpassword.data, stored_user['salt'], stored_user['hashed']):
            user = User(form.loginemail.data)
            login_user(user, remember=True)
            return redirect(url_for('lookup'))
        form.loginemail.errors.append("Email or password invalid")
    return render_template("home.html", loginform=form, registrationform=RegistrationForm())


@app.route("/register", methods=["POST"])
def register():
    form = RegistrationForm(request.form)
    if form.validate():
        if DB.get_user(form.email.data):
            form.email.errors.append("Email address already registered")
            return render_template("home.html", loginform=LoginForm(), registrationform=form)
        salt = PH.get_salt()
        hashed = PH.get_hash(form.password2.data + salt)
        DB.add_user(form.email.data, salt, hashed)
        return render_template("home.html", loginform=LoginForm(), registrationform=form, onloadmessage="Registration successful. Please log in.")
    return render_template("home.html", loginform=LoginForm(), registrationform=form)


@app.route("/logout")
def logout():
    logout_user()
    return redirect(url_for("home"))


@app.route("/dashboard")
@login_required
def dashboard():
    nownow = datetime.datetime.now().strftime('%Y-%m-%d')
    lattice_cnt = DB.get_site_count({"site.tower_type":"Lattice"})
    monopole_cnt = DB.get_site_count({"site.tower_type":"Monopole"})
    monolattice_cnt = DB.get_site_count({"site.tower_type":"Mono-Lattice"})
    total_cnt = DB.get_site_count({})
    context = {
        'features': [
            {'description': 'Lattice', 'value': lattice_cnt},
            {'description': 'Monopole', 'value': monopole_cnt},
            {'description': 'Mono-Lattice', 'value': monolattice_cnt},
            {'description': 'Total', 'value': total_cnt},
        ], 'date_now': nownow
    }
    return render_template("dashboard.html", **context)


@app.route("/dashboard/resolve")
@login_required
def dashboard_resolve():
    request_id = request.args.get("request_id")
    DB.delete_request(request_id)
    return redirect(url_for('dashboard'))


@app.route("/lookup")
@login_required
def lookup():
    return render_template("lookup.html", searchform=SearchForm())


@app.route("/lookup/search", methods=["POST"])
@login_required
def lookup_search():
    form = SearchForm(request.form)
    print form.data
    if form.validate():
        query = make_query(site_name=form.site_name.data,
                            site_number=form.site_number.data,
                            region=form.region.data,
                            type=form.tower_type.data,
                            date_of_inspection=form.date_of_inspection.data
                            )
        sites = DB.find_sites(query)
        print "Length for sites : " + str(len(sites))
        site = sites[0]
        print site
        return render_template("lookup.html", searchform=form, sites=sites)
    return redirect(url_for('lookup'))


@app.route("/lookup/showsite", methods=["POST", "GET"])
@login_required
def lookup_showsite():
    siteid = request.args.get("siteid")
    if DB.check_manual_exists(siteid):
        print "Inside /lookup/showsite - and DB.manual_exists is True"
    else:
        print "Inside /lookup/showsite - and DB.manual_exists is False"
        context = {
            'site.manual.updated': 'False',
            'site.manual.site_classification': "Fat Charlie",
            'site.manual.site_release_date': "",
            'site.manual.as_built_available': "",
            'site.manual.fault_description': "",
            'site.manual.mast_engineer': "",
            'site.manual.mast_upgraded': "",
            'site.manual.mast_upgraded_date': "",
            'site.manual.capacity_top': "",
            'site.manual.capacity_10_from_top': "",
            'site.manual.update_date': ""
        }
        DB.update_site_manual(siteid, context)
    sites = list(DB.show_site(siteid))
    print "Length for sites : " + str(len(sites))
    site = sites[0]
    print site
    return render_template("showsite.html", updateform=UpdateForm(), site=site)


@app.route("/lookup/update", methods=["POST", "GET"])
@login_required
def lookup_updatesite():
    # siteid = request.args.get("siteid")
    form = UpdateForm(request.form)
    siteid = form.siteid.data
    print form.data
    print "Inside lookup/update - about to validate form"
    print siteid
    print form.site_classification.data
    nownow = datetime.datetime.now().strftime('%Y-%m-%d')
    if form.validate():
        context = {
            'updated': 'True',
            'site_classification': form.site_classification.data,
            'site_release_date': form.site_release_date.data,
            'as_built_available': form.as_built_available.data,
            'fault_description': form.fault_description.data,
            'mast_engineer': form.mast_engineer.data,
            'mast_upgraded': form.mast_upgraded.data,
            'mast_upgraded_date': form.mast_upgraded_date.data,
            'capacity_top': form.capacity_top.data,
            'capacity_10_from_top': form.capacity_10_from_top.data,
            'date_now': nownow
        }
        print "Show context"
        print context
        DB.update_site_manual(siteid, context)
        print "Update site with context"
        sites = list(DB.show_site(siteid))
        print "Length for sites : " + str(len(sites))
        site = sites[0]
        print "Show site after fetch"
        print site
        # return render_template("showsite.html", updateform=form, site=site)
    return redirect(url_for('lookup'))



@app.route("/output")
@login_required
def output():
    return render_template("output.html", searchform=SearchForm())


@app.route("/output/search", methods=["POST"])
@login_required
def output_search():
    form = SearchForm(request.form)
    if form.validate():
        query = make_query(
            site_name=form.site_name.data,
            site_number=form.site_number.data,
            region=form.region.data,
            type=form.tower_type.data,
            date_of_inspection=form.date_of_inspection.data
            )

        sites = list(DB.find_sites(query))

        main_filename = save_to_workbook(sites, output_columns_main, "sites.xlsx")

        # loading_filename = save_to_workbook(sites, output_columns_loadings, "loadings.xlsx")

        return render_template('download.html')
    else:
        print "Form failed to validate"
    return redirect(url_for('output'))


@app.route("/output/download", methods=["POST"])
@login_required
def output_download():
    filename = "sites.xlsx"
    parentddir = os.path.abspath(os.path.dirname(__file__))
    send_from_directory(directory=parentddir, filename=filename)

    return redirect(url_for('output'))


@app.route('/download-file/')
def return_files_tut():
    try:
        parentddir = os.path.abspath(os.path.dirname(__file__))
        filename = 'sites.xlsx'
        full_path = os.path.join(parentddir,filename)
        return send_file(full_path, attachment_filename=filename)
    except Exception as e:
        return str(e)

# Create a workbook and add a worksheet.
# workbook = load_workbook('sites.xlsx')
def save_to_workbook(sites, output_columns, output_file):

    workbook = Workbook()

    # worksheet = workbook.active
    worksheet = workbook.create_sheet(title="Data", index=0)

    print output_columns

    # Start from the first cell. Rows and columns are zero indexed.
    row = 1
    for tag, value in output_columns.iteritems():
        _ = worksheet.cell(row=row, column=value, value=tag)

    row = 2
    for site in sites:

        print "Site data : " + str(site)

        # Iterate over the data and write it out row by row.
        for tag, value in output_columns.iteritems():
            col = output_columns[tag]     # get the column from the dictionary
            value = get_site_value(site,tag)
            # print "Tag : " + tag + " Value: " + str(value)
            _ = worksheet.cell(row=row, column=col, value=value)
            col += 1

        row += 1
        col = 1

    # Save the file
    output_filename = output_file
    workbook.save(output_filename)
    print "Output file : " + output_filename + " saved"
    return output_filename


def get_site_value(site, tag):
    value = None
    mongo_str = output_lookup[tag]

    if mongo_str:
        args = mongo_str.split(".")
        value = dict_digger.dig(site, *args)

    return value


def make_query(site_name=None, site_number=None, region=None, type=None, date_of_inspection=None):
    # all queries begin with something common, which may be an empty dict, but here's an example
    query = {}

    if site_name:
        clean_str = site_name.strip(" ")
        regexp = re.compile(clean_str, re.IGNORECASE)
        query['site.site_group.site_name'] = regexp
    if site_number:
        clean_str = site_number.strip(" ")
        regexp = re.compile(clean_str, re.IGNORECASE)
        query['site.site_group.site_number'] = regexp
    if region:
        query['site.site_group.region'] = region
    if type:
        query['site.tower_type'] = type
    if date_of_inspection:
        query['site.date_of_inspection'] = date_of_inspection
    # etc...

    return query


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='10.240.221.76', port=port, debug=True)
